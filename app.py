from flask import Flask, render_template, request, redirect, session
import os
import fitz  # PyMuPDF for Engineering PDF Drawing Scanning
import mysql.connector
from werkzeug.utils import secure_filename
import re
from datetime import datetime
from openpyxl import load_workbook
from urllib.parse import quote
from zipfile import ZipFile
import xml.etree.ElementTree as ET

# Load environment variables safely if configured
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from vendor_checker import check_vendor
from detailed_engine import get_detail
from human_reviewer import review as generate_human_review

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "ntpcsecret")

UPLOAD_FOLDER = "uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

uploaded_pdfs = []

# ================= SECURE MYSQL CONNECTION =================
try:
    db = mysql.connector.connect(
        host=os.getenv("DB_HOST", "localhost"),
        user=os.getenv("DB_USER", "root"),
        password=os.getenv("DB_PASSWORD", "Arjun@0901"),
        database=os.getenv("DB_NAME", "Project")
    )
    cursor = db.cursor(buffered=True)
    print("Database Connected Successfully")
except Exception as e:
    print("DATABASE ERROR")
    print(e)

def initialize_database_schemas():
    """Ensures all original and tracking tables are initialized securely in MySQL"""
    try:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INT AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(100),
                mobile VARCHAR(15),
                password VARCHAR(100),
                role VARCHAR(50),
                approved INT DEFAULT 0
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS vendor_uploads (
                id INT AUTO_INCREMENT PRIMARY KEY,
                vendor_name VARCHAR(100),
                uploaded_by INT,
                pdf_name VARCHAR(255),
                file_path VARCHAR(255),
                status VARCHAR(50),
                accuracy INT,
                upload_date DATETIME
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS progress_uploads (
                id INT AUTO_INCREMENT PRIMARY KEY,
                vendor_name VARCHAR(100),
                pdf_name VARCHAR(255),
                upload_date DATE
            )
        """)
        # Persistent Analytics Structure: Resolves the missing SQL values issue
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS running_progress_metrics (
                id INT AUTO_INCREMENT PRIMARY KEY,
                vendor_name VARCHAR(100),
                filename VARCHAR(255),
                activity_name VARCHAR(255),
                serial_no VARCHAR(50),
                uom VARCHAR(50),
                scope DOUBLE DEFAULT 0.0,
                previous_day_cumm DOUBLE DEFAULT 0.0,
                today_progress DOUBLE DEFAULT 0.0,
                cumulative_completion DOUBLE DEFAULT 0.0,
                to_be_completed DOUBLE DEFAULT 0.0,
                completion_percentage DOUBLE DEFAULT 0.0,
                remarks TEXT,
                logged_date DATE
            )
        """)
        running_progress_columns = {
            "serial_no": "VARCHAR(50)",
            "uom": "VARCHAR(50)",
            "previous_day_cumm": "DOUBLE DEFAULT 0.0",
            "cumulative_completion": "DOUBLE DEFAULT 0.0",
            "to_be_completed": "DOUBLE DEFAULT 0.0",
            "remarks": "TEXT"
        }
        for column_name, column_type in running_progress_columns.items():
            cursor.execute("SHOW COLUMNS FROM running_progress_metrics LIKE %s", (column_name,))
            if not cursor.fetchone():
                cursor.execute(f"ALTER TABLE running_progress_metrics ADD COLUMN {column_name} {column_type}")
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS ai_learning (
                id INT AUTO_INCREMENT PRIMARY KEY,
                document_name VARCHAR(255),
                upload_date DATETIME,
                extracted_text LONGTEXT,
                document_type VARCHAR(100),
                uploaded_by VARCHAR(100)
            )
        """)
        ai_learning_columns = {
            "document_name": "VARCHAR(255)",
            "upload_date": "DATETIME",
            "extracted_text": "LONGTEXT",
            "document_type": "VARCHAR(100)",
            "uploaded_by": "VARCHAR(100)"
        }
        for column_name, column_type in ai_learning_columns.items():
            cursor.execute("SHOW COLUMNS FROM ai_learning LIKE %s", (column_name,))
            if not cursor.fetchone():
                cursor.execute(f"ALTER TABLE ai_learning ADD COLUMN {column_name} {column_type}")
        db.commit()
    except Exception as e:
        print("Schema setup warning context:", e)

initialize_database_schemas()

# ================= CORE DATA MATRIX PARSER PIPELINE =================
def process_dpr_file_data(vendor_name, filename, filepath, report_date):
    """Parses NTPC DPR Excel layouts and keeps the marked review columns intact."""
    try:
        cursor.execute("""
            DELETE FROM running_progress_metrics
            WHERE vendor_name=%s AND filename=%s AND logged_date=%s
        """, (vendor_name, filename, report_date))
        db.commit()

        wb = load_workbook(filepath, data_only=True)
        sheet_target = None
        for s_name in ["1200 MW Summary", "1200MW Summary", "Summary"]:
            if s_name in wb.sheetnames:
                sheet_target = wb[s_name]
                break
        if sheet_target is None:
            sheet_target = wb.active

        def clean_label(value):
            return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip().lower()

        def safe_number(value):
            if value in (None, ""):
                return 0.0
            if isinstance(value, str):
                value = value.strip().replace(",", "").replace("%", "")
                if value in ("", "-"):
                    return 0.0
            try:
                return float(value)
            except (TypeError, ValueError):
                return 0.0

        def find_column(header_row, checks):
            for c in range(1, sheet_target.max_column + 1):
                label = clean_label(sheet_target.cell(row=header_row, column=c).value)
                if all(check in label for check in checks):
                    return c
            return None

        header_cells = []
        for r in range(1, min(sheet_target.max_row, 20) + 1):
            for c in range(1, sheet_target.max_column + 1):
                if clean_label(sheet_target.cell(row=r, column=c).value) == "activity":
                    header_cells.append((r, c))

        if not header_cells:
            header_cells = [(3, 2)]

        for header_row, activity_col in header_cells:
            sr_col = find_column(header_row, ["sr"]) or (activity_col - 1 if activity_col > 1 else None)
            uom_col = find_column(header_row, ["uom"]) or activity_col + 1
            scope_col = find_column(header_row, ["scope"]) or activity_col + 2
            previous_col = find_column(header_row, ["day", "cumm"]) or find_column(header_row, ["pre", "cumulative"]) or activity_col + 3
            today_col = find_column(header_row, ["today", "progress"]) or activity_col + 4
            cumulative_col = find_column(header_row, ["cumulative", "completion"]) or find_column(header_row, ["cumulative", "progress"]) or activity_col + 5
            to_be_col = find_column(header_row, ["to be"]) or activity_col + 6
            completion_col = find_column(header_row, ["%"]) or find_column(header_row, ["of", "completion"]) or activity_col + 7
            remarks_col = find_column(header_row, ["remarks"]) or activity_col + 8

            for r in range(header_row + 1, sheet_target.max_row + 1):
                try:
                    act_val = sheet_target.cell(row=r, column=activity_col).value
                    activity_lbl = str(act_val or "").strip()
                    if not activity_lbl or "PV Plant" in activity_lbl:
                        continue

                    if activity_lbl in ["Activity", "Description", "UoM", "Scope", "Remarks"]:
                        continue

                    serial_no = sheet_target.cell(row=r, column=sr_col).value if sr_col else ""
                    uom = sheet_target.cell(row=r, column=uom_col).value
                    raw_scope = sheet_target.cell(row=r, column=scope_col).value
                    raw_previous = sheet_target.cell(row=r, column=previous_col).value
                    raw_today = sheet_target.cell(row=r, column=today_col).value
                    raw_cumulative = sheet_target.cell(row=r, column=cumulative_col).value
                    raw_to_be_completed = sheet_target.cell(row=r, column=to_be_col).value
                    raw_comp = sheet_target.cell(row=r, column=completion_col).value
                    remarks = sheet_target.cell(row=r, column=remarks_col).value

                    scope_val = safe_number(raw_scope)
                    previous_val = safe_number(raw_previous)
                    today_val = safe_number(raw_today)
                    cumulative_val = safe_number(raw_cumulative)
                    to_be_completed_val = safe_number(raw_to_be_completed)
                    comp_percentage = safe_number(raw_comp)
                    
                    if comp_percentage and comp_percentage <= 1.0:
                        comp_percentage = comp_percentage * 100.0
                    if comp_percentage > 100.0: comp_percentage = 100.0

                    if not any([scope_val, previous_val, today_val, cumulative_val, to_be_completed_val, comp_percentage]):
                        continue

                    cursor.execute("""
                        INSERT INTO running_progress_metrics 
                        (vendor_name, filename, activity_name, serial_no, uom, scope, previous_day_cumm,
                         today_progress, cumulative_completion, to_be_completed, completion_percentage,
                         remarks, logged_date)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        vendor_name, filename, activity_lbl, str(serial_no or ""), str(uom or ""),
                        scope_val, previous_val, today_val, cumulative_val, to_be_completed_val,
                        comp_percentage, str(remarks or ""), report_date
                    ))
                except Exception:
                    continue
        db.commit()
    except Exception as e:
        print(f"Deep tracking matrix parse fault for {filename}: {e}")

def refresh_progress_metrics_from_saved_files():
    """Rebuilds DPR metrics from saved Excel uploads so older partial rows do not keep stale graph data."""
    try:
        cursor.execute("SELECT vendor_name, pdf_name, upload_date FROM progress_uploads ORDER BY upload_date ASC")
        for vendor_name, filename, report_date in cursor.fetchall():
            if not filename:
                continue

            ext = os.path.splitext(filename)[1].lower()
            if ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
                continue

            filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            if not os.path.exists(filepath):
                continue

            process_dpr_file_data(vendor_name or "0", filename, filepath, report_date)
    except Exception as e:
        print("Progress metrics refresh warning:", e)

def find_learned_references(source_text, limit=5):
    words = re.findall(r"[A-Za-z][A-Za-z0-9_-]{4,}", source_text.lower())
    stop_words = {
        "shall", "drawing", "ntpc", "section", "layout", "vendor", "project",
        "document", "engineering", "review", "system", "solar", "plant"
    }
    keywords = []
    for word in words:
        if word not in stop_words and word not in keywords:
            keywords.append(word)
        if len(keywords) >= 8:
            break

    if not keywords:
        return []

    clauses = " OR ".join(["extracted_text LIKE %s"] * len(keywords))
    params = [f"%{keyword}%" for keyword in keywords]

    try:
        cursor.execute(f"""
            SELECT document_name, document_type, LEFT(extracted_text, 220)
            FROM ai_learning
            WHERE {clauses}
            ORDER BY upload_date DESC
            LIMIT {limit}
        """, params)
        return cursor.fetchall()
    except Exception as e:
        print("Learned reference lookup warning:", e)
        return []


# ================= HOME INTERFACES =================
@app.route("/")
def firstpage():
    return redirect("/login")

@app.route("/home")
def home():
    if "username" not in session:
        return redirect("/login")
    try:
        cursor.execute("SELECT COUNT(*) FROM vendor_uploads")
        total_drawings = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM vendor_uploads WHERE status='Approved'")
        approved_count = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM vendor_uploads WHERE status='Pending'")
        pending_count = cursor.fetchone()[0]
    except:
        total_drawings = approved_count = pending_count = 0

    return render_template(
        "index.html",
        total_drawings=total_drawings,
        approved_count=approved_count,
        pending_count=pending_count,
        ai_confidence="94%"
    )


# ================= USER AUTHENTICATION LAYER =================
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        cursor.execute("SELECT * FROM users WHERE username=%s", (username,))
        user = cursor.fetchone()

        if user:
            dbpassword = user[3]
            role = user[4]
            approved = user[5]

            if password == dbpassword:
                session["user_id"] = user[0]
                session["username"] = username
                session["role"] = role

                if role == "prime_admin":
                    return redirect("/prime_dashboard")
                elif role == "admin":
                    if approved == 0:
                        return """
                        <script>
                        alert("Waiting Prime Admin Approval");
                        window.location="/login";
                        </script>
                        """
                    return redirect("/home")
                elif role == "vendor":
                    return redirect("/vendor_dashboard")

        return """
        <script>
        alert("Wrong username or password");
        window.location="/login";
        </script>
        """
    return render_template("login.html")

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form["username"].strip()
        email = request.form["email"].strip()
        mobile = request.form["mobile"].strip()
        password = request.form["password"]
        role = request.form["role"]

        if len(username) < 4:
            return '<script>alert("Username minimum 4 characters"); window.location="/register";</script>'
        if not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
            return '<script>alert("Enter valid email"); window.location="/register";</script>'
        if not re.match(r'^[6-9]\d{9}$', mobile):
            return '<script>alert("Enter valid mobile number"); window.location="/register";</script>'
        if len(password) < 6:
            return '<script>alert("Password minimum 6 chars"); window.location="/register";</script>'

        approval = 0 if role == "admin" else 1

        try:
            cursor.execute("SELECT * FROM users WHERE username=%s", (username,))
            if cursor.fetchone():
                return '<script>alert("Username already exists"); window.location="/register";</script>'

            cursor.execute("""
                INSERT INTO users (username, mobile, password, role, approved)
                VALUES (%s, %s, %s, %s, %s)
            """, (username, mobile, password, role, approval))
            db.commit()

            return '<script>alert("Registration Successful"); window.location="/login";</script>'
        except Exception as e:
            return f'<script>alert("{str(e)}"); window.location="/register";</script>'
    return render_template("register.html")


# ================= PRIME ADMIN CONTROLS =================
@app.route("/prime_dashboard")
def prime_dashboard():
    if "username" not in session or session["role"] != "prime_admin":
        return redirect("/login")

    cursor.execute("SELECT * FROM users WHERE role='admin' AND approved=0")
    admins = cursor.fetchall()
    return render_template("prime_dashboard.html", admins=admins)

@app.route("/approve_admin/<username>")
def approve_admin(username):
    cursor.execute("UPDATE users SET approved=1 WHERE username=%s", (username,))
    db.commit()
    return redirect("/prime_dashboard")


# ================= ENGINEERING DRAWING REVIEW CRADLE =================
@app.route("/upload", methods=["POST"])
def upload():
    global uploaded_pdfs
    if "username" not in session:
        return redirect("/login")

    uploaded_pdfs = []
    files = request.files.getlist("vendor_pdf")

    if len(files) == 0 or files[0].filename == "":
        return '<script>alert("No file selected"); window.location="/vendor_dashboard";</script>'

    vendor = session["username"]
    user_id = session["user_id"]

    for file in files:
        if file.filename == "": continue
        filename = secure_filename(file.filename)
        path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        file.save(path)
        uploaded_pdfs.append(path)

        try:
            doc = fitz.open(path)
            text = "".join([page.get_text() for page in doc]).strip()
            total_chars, total_words = len(text), len(text.split())

            if total_chars < 30:
                vendor_accuracy = 0
            else:
                result = check_vendor(path)
                compliance = int((result["found"] / max(result.get("total", 1), 1)) * 100)
                historical_score = min(total_words // 8, 100)
                human_score = min(total_chars // 50, 100)
                vendor_accuracy = int((compliance * .60) + (historical_score * .25) + (human_score * .15))

                if vendor_accuracy > 100: vendor_accuracy = 100

            cursor.execute("""
                INSERT INTO vendor_uploads (vendor_name, uploaded_by, pdf_name, file_path, status, accuracy, upload_date)
                VALUES (%s, %s, %s, %s, 'Pending', %s, NOW())
            """, (vendor, user_id, filename, path, vendor_accuracy))
            db.commit()
        except Exception as e:
            print("UPLOAD ERROR:", e)

    return '<script>alert("Drawing Uploaded Successfully"); window.location="/vendor_dashboard";</script>'

@app.route("/pending")
def pending():
    if "username" not in session: return redirect("/login")
    cursor.execute("SELECT id, vendor_name, uploaded_by, pdf_name, file_path, status, accuracy, upload_date FROM vendor_uploads WHERE status='Pending' ORDER BY id DESC")
    files = cursor.fetchall()
    return render_template("pending.html", files=files)

@app.route("/approved")
def approved_drawings():
    if "username" not in session: return redirect("/login")
    cursor.execute("SELECT id, vendor_name, uploaded_by, pdf_name, file_path, status, accuracy, upload_date FROM vendor_uploads WHERE status='Approved' ORDER BY id DESC")
    files = cursor.fetchall()
    return render_template("approved.html", files=files)

@app.route("/rejected")
def rejected_drawings():
    if "username" not in session: return redirect("/login")
    cursor.execute("SELECT id, vendor_name, uploaded_by, pdf_name, file_path, status, accuracy, upload_date FROM vendor_uploads WHERE status='Rejected' ORDER BY id DESC")
    data = cursor.fetchall()
    return render_template("rejected.html", files=data)


# ================= DEEP WORK SCANNING & ANALYSIS CONTROLLERS =================
@app.route("/analyze/<int:id>")
def analyze(id):
    cursor.execute("SELECT file_path FROM vendor_uploads WHERE id=%s", (id,))
    file = cursor.fetchone()
    if file:
        uploaded_pdfs.clear()
        uploaded_pdfs.append(file[0])
        return redirect("/scan")
    return redirect("/pending")

@app.route("/scan")
def scan():
    path = uploaded_pdfs[0]
    result = check_vendor(path)
    doc = fitz.open(path)
    text = "".join([page.get_text() for page in doc]).strip()
    total_chars, total_words = len(text), len(text.split())

    if total_chars < 30:
        compliance = vendor_accuracy = historical_score = human_score = 0
    else:
        compliance = int((result["found"] / max(result.get("total", 1), 1)) * 100)
        historical_score = min(total_words // 8, 100)
        human_score = min(total_chars // 50, 100)
        vendor_accuracy = int((compliance * .60) + (historical_score * .25) + (human_score * .15))

    learned_refs = find_learned_references(text)
    if learned_refs:
        learned_review = "Learned Knowledge References:\n"
        for ref in learned_refs:
            learned_review += f"- {ref[0]} ({ref[1]}): {ref[2]}\n"
    else:
        learned_review = "Learned Knowledge References:\n- No learned knowledge match found yet. Upload standards/manuals in AI Learning Centre."

    reviewer_summary = generate_human_review(
        result["engineering"],
        result["layout"],
        result["foundation"],
        result["electrical"],
        result["structural"],
        result["safety"],
        vendor_accuracy
    )

    return render_template(
        "dashboard.html",
        engineering=result["engineering"], layout=result["layout"], foundation=result["foundation"],
        electrical=result["electrical"], structural=result["structural"], safety=result["safety"],
        found=result["found"], total=result.get("total", 0), compliance=compliance, vendor_accuracy=vendor_accuracy,
        historical_score=historical_score, human_score=human_score,
        status="Pending AI Review", ai_review=f"{reviewer_summary}\n\n{learned_review}"
    )

@app.route("/approve_vendor/<int:id>")
def approve_vendor(id):
    cursor.execute("UPDATE vendor_uploads SET status='Approved' WHERE id=%s", (id,))
    db.commit()
    return redirect("/pending")

@app.route("/approved_dashboard/<int:id>")
def approved_dashboard(id):
    try:
        cursor.execute("UPDATE vendor_uploads SET status='Approved' WHERE id=%s", (id,))
        db.commit()
    except Exception as e:
        print("APPROVE ERROR:", e)
    return redirect("/approved")

@app.route("/reject_vendor/<int:id>")
def reject_vendor(id):
    cursor.execute("UPDATE vendor_uploads SET status='Rejected' WHERE id=%s", (id,))
    db.commit()
    return redirect("/pending")

@app.route("/reject", methods=["POST"])
def reject():
    cursor.execute("UPDATE vendor_uploads SET status='Rejected' ORDER BY id DESC LIMIT 1")
    db.commit()
    return '<script>alert("Drawing Rejected"); window.location="/rejected";</script>'

def delete_vendor_upload_record(record_id):
    cursor.execute("SELECT file_path FROM vendor_uploads WHERE id=%s", (record_id,))
    row = cursor.fetchone()
    if not row:
        return False, "Drawing record not found"

    file_path = row[0]
    cursor.execute("DELETE FROM vendor_uploads WHERE id=%s", (record_id,))
    db.commit()

    if file_path and os.path.exists(file_path):
        try:
            os.remove(file_path)
        except OSError as e:
            return True, f"Database record deleted, but file could not be removed: {e}"

    return True, "Drawing deleted successfully"

@app.route("/delete_vendor/<int:id>")
def delete_vendor(id):
    source_page = request.args.get("from", "pending")
    redirect_map = {
        "pending": "/pending",
        "approved": "/approved",
        "rejected": "/rejected",
        "vendor": "/vendor_dashboard",
        "home": "/home"
    }
    success, message = delete_vendor_upload_record(id)
    target = redirect_map.get(source_page, "/pending")
    status = "success" if success else "error"
    return redirect(f"{target}?{status}={quote(message)}")

@app.route("/delete_vendor/<path:filename>")
def delete_vendor_by_filename(filename):
    source_page = request.args.get("from", "pending")
    cursor.execute("SELECT id FROM vendor_uploads WHERE pdf_name=%s ORDER BY id DESC LIMIT 1", (filename,))
    row = cursor.fetchone()
    if not row:
        return redirect(f"/{source_page}?error={quote('Drawing record not found')}")
    success, message = delete_vendor_upload_record(row[0])
    status = "success" if success else "error"
    return redirect(f"/{source_page}?{status}={quote(message)}")

@app.route("/delete/<int:id>")
def delete(id):
    success, message = delete_vendor_upload_record(id)
    status = "success" if success else "error"
    return redirect(f"/vendor_dashboard?{status}={quote(message)}")

@app.route("/detail")
def detail():
    issue = request.args.get("issue", "Engineering Drawing Review")
    detail_data = get_detail(issue)
    analysis = f"""
Section: {detail_data['section']}
Clause: {detail_data['clause']}
Severity: {detail_data['severity']}

Requirement:
{detail_data['requirement']}

Engineering Impact:
{detail_data['engineering_impact']}

AI Reasoning:
{detail_data['ai_reasoning']}

Recommendation:
{detail_data['recommendation']}

Suggested Status:
{detail_data['suggested_status']}
"""
    return render_template("detail.html", issue=issue, analysis=analysis)

@app.route("/vendor_dashboard")
def vendor_dashboard():
    if "username" not in session: return redirect("/login")
    vendor, user_id = session["username"], session["user_id"]
    cursor.execute("SELECT id, vendor_name, uploaded_by, pdf_name, file_path, status, accuracy, upload_date FROM vendor_uploads WHERE uploaded_by=%s ORDER BY id DESC", (user_id,))
    data = []
    for row in cursor.fetchall():
        data.append({
            "id": row[0],
            "vendor_name": row[1],
            "uploaded_by": row[2],
            "pdf_name": row[3],
            "file_path": row[4],
            "status": row[5],
            "accuracy": row[6] or 0,
            "upload_date": row[7]
        })
    return render_template("vendor_dashboard.html", vendor=vendor, files=data)

@app.route("/vendor_progress")
def vendor_progress():
    if "username" not in session: return redirect("/login")
    return render_template("vendor_progress.html")

def extract_text_from_pdf(filepath):
    doc = fitz.open(filepath)
    return "\n".join(page.get_text() for page in doc).strip()

def extract_text_from_docx(filepath):
    try:
        with ZipFile(filepath) as docx_zip:
            xml_data = docx_zip.read("word/document.xml")
        root = ET.fromstring(xml_data)
        namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        paragraphs = []
        for paragraph in root.findall(".//w:p", namespace):
            words = [node.text for node in paragraph.findall(".//w:t", namespace) if node.text]
            if words:
                paragraphs.append(" ".join(words))
        return "\n".join(paragraphs).strip()
    except Exception as e:
        print("DOCX extraction warning:", e)
        return ""

def extract_text_from_excel(filepath):
    try:
        workbook = load_workbook(filepath, data_only=True, read_only=True)
        rows = []
        for sheet in workbook.worksheets:
            rows.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value not in (None, "")]
                if values:
                    rows.append(" | ".join(values))
        return "\n".join(rows).strip()
    except Exception as e:
        print("Excel extraction warning:", e)
        return ""

def extract_learning_text(filepath, filename):
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".pdf":
        return extract_text_from_pdf(filepath)
    if ext == ".docx":
        return extract_text_from_docx(filepath)
    if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        return extract_text_from_excel(filepath)
    if ext in [".txt", ".csv"]:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as file:
            return file.read().strip()
    return ""

def classify_learning_document(filename, selected_type):
    if selected_type:
        return selected_type
    ext = os.path.splitext(filename)[1].lower().replace(".", "").upper()
    if ext in ["XLSX", "XLSM", "CSV"]:
        return "Excel / DPR / Data"
    if ext == "DOCX":
        return "Engineering Manual"
    if "standard" in filename.lower():
        return "Standard"
    if "guideline" in filename.lower():
        return "NTPC Guideline"
    if "drawing" in filename.lower():
        return "Approved Drawing"
    return "Knowledge Document"

@app.route("/learning", methods=["GET", "POST"])
def learning():
    if "username" not in session: return redirect("/login")

    message, error = request.args.get("success"), request.args.get("error")

    if request.method == "POST":
        learning_file = request.files.get("learning_file")
        document_type = request.form.get("document_type", "").strip()

        if not learning_file or learning_file.filename == "":
            return redirect("/learning?error=No%20file%20selected")

        filename = secure_filename(learning_file.filename)
        ext = os.path.splitext(filename)[1].lower()
        if ext not in [".pdf", ".docx", ".xlsx", ".xlsm", ".xltx", ".xltm", ".txt", ".csv"]:
            return redirect("/learning?error=Unsupported%20knowledge%20file%20type")

        filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        learning_file.save(filepath)

        extracted_text = extract_learning_text(filepath, filename)
        if not extracted_text:
            return redirect("/learning?error=No%20text%20could%20be%20extracted%20from%20this%20document")

        try:
            cursor.execute("""
                INSERT INTO ai_learning (document_name, upload_date, extracted_text, document_type, uploaded_by)
                VALUES (%s, NOW(), %s, %s, %s)
            """, (
                filename,
                extracted_text,
                classify_learning_document(filename, document_type),
                session["username"]
            ))
            db.commit()
            return redirect("/learning?success=Knowledge%20document%20learned%20successfully")
        except Exception as e:
            print("AI learning insert error:", e)
            return redirect("/learning?error=Knowledge%20document%20could%20not%20be%20saved")

    search_query = request.args.get("q", "").strip()
    if search_query:
        like_query = f"%{search_query}%"
        cursor.execute("""
            SELECT document_name, document_type, uploaded_by, upload_date, LEFT(extracted_text, 600)
            FROM ai_learning
            WHERE extracted_text LIKE %s OR document_name LIKE %s OR document_type LIKE %s
            ORDER BY upload_date DESC
            LIMIT 25
        """, (like_query, like_query, like_query))
        search_results = cursor.fetchall()
    else:
        search_results = []

    cursor.execute("SELECT COUNT(*), COALESCE(SUM(CHAR_LENGTH(extracted_text)), 0), MAX(upload_date) FROM ai_learning")
    stats_row = cursor.fetchone()

    cursor.execute("SELECT COUNT(DISTINCT document_type) FROM ai_learning")
    standards_count = cursor.fetchone()[0] or 0

    cursor.execute("""
        SELECT id, document_name, document_type, uploaded_by, upload_date
        FROM ai_learning
        ORDER BY upload_date DESC
        LIMIT 50
    """)
    learned_documents = cursor.fetchall()

    stats = {
        "total_documents": stats_row[0] or 0,
        "total_standards": standards_count,
        "knowledge_size": f"{int(stats_row[1] or 0):,} chars",
        "last_learning_date": stats_row[2].strftime("%d-%b-%Y %H:%M") if stats_row and stats_row[2] else "0"
    }

    return render_template(
        "learning.html",
        stats=stats,
        learned_documents=learned_documents,
        search_query=search_query,
        search_results=search_results,
        success_message=message,
        error_message=error
    )


# ================= HIGH-FIDELITY TELEMETRY INJECTION SYSTEM =================
@app.route("/upload_progress", methods=["POST"])
def upload_progress():
    if "username" not in session: return redirect("/login")
    files = request.files.getlist("progress_file")
    if not files or files[0].filename == "": files = request.files.getlist("progress_pdf")

    uploaded_by_vendor = session["username"]

    for file in files:
        if file.filename == "": continue
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        file.save(filepath)

        report_date = None
        found = re.search(r'(\d{2})-(\d{2})-(\d{2})', filename)
        if found:
            day, month, year = found.group(1), found.group(2), "20" + found.group(3)
            report_date = f"{year}-{month}-{day}"
        
        if not report_date:
            try:
                wb = load_workbook(filepath, data_only=True)
                sheet = wb.active
                cell_val = sheet["D2"].value
                if isinstance(cell_val, datetime): report_date = cell_val.strftime("%Y-%m-%d")
            except:
                pass
        
        if not report_date: report_date = datetime.now().strftime("%Y-%m-%d")

        try:
            cursor.execute("INSERT INTO progress_uploads (vendor_name, pdf_name, upload_date) VALUES (%s, %s, %s)", (uploaded_by_vendor, filename, report_date))
            db.commit()
            process_dpr_file_data(uploaded_by_vendor, filename, filepath, report_date)
        except Exception as database_err:
            print("Error in upload_progress module:", database_err)

    return redirect("/progress_review")


@app.route("/progress_review")
def progress_review():
    if "username" not in session: return redirect("/login")

    def format_dpr_date(value):
        if not value:
            return "0"
        try:
            return value.strftime("%d-%b-%Y")
        except AttributeError:
            return str(value)

    def build_activity_chart(activity_pattern, minimum_scope=0):
        cursor.execute("""
            SELECT
                logged_date,
                COALESCE(MAX(scope), 0),
                COALESCE(MAX(
                    CASE
                        WHEN previous_day_cumm = 0 AND cumulative_completion = 0 AND completion_percentage > 0 AND scope > 0
                        THEN GREATEST((scope * completion_percentage / 100) - today_progress, 0)
                        ELSE previous_day_cumm
                    END
                ), 0),
                COALESCE(SUM(today_progress), 0),
                COALESCE(MAX(
                    CASE
                        WHEN cumulative_completion = 0 AND completion_percentage > 0 AND scope > 0
                        THEN scope * completion_percentage / 100
                        ELSE cumulative_completion
                    END
                ), 0),
                COALESCE(MAX(
                    CASE
                        WHEN to_be_completed = 0 AND completion_percentage > 0 AND scope > 0
                        THEN GREATEST(scope - (scope * completion_percentage / 100), 0)
                        ELSE to_be_completed
                    END
                ), 0)
            FROM running_progress_metrics
            WHERE LOWER(TRIM(activity_name)) LIKE %s
                AND COALESCE(scope, 0) >= %s
            GROUP BY logged_date
            ORDER BY logged_date ASC
        """, (activity_pattern, minimum_scope))
        chart_rows = cursor.fetchall()

        activity_chart = {
            "dates": [],
            "scope": [],
            "previous": [],
            "today": [],
            "cumulative": [],
            "to_be_completed": []
        }

        for row in chart_rows:
            activity_chart["dates"].append(format_dpr_date(row[0]))
            activity_chart["scope"].append(round(float(row[1] or 0.0), 2))
            activity_chart["previous"].append(round(float(row[2] or 0.0), 2))
            activity_chart["today"].append(round(float(row[3] or 0.0), 2))
            activity_chart["cumulative"].append(round(float(row[4] or 0.0), 2))
            activity_chart["to_be_completed"].append(round(float(row[5] or 0.0), 2))

        return activity_chart

    pile_casting_chart = build_activity_chart("%pile%casting%", 50000)
    module_installation_chart = build_activity_chart("%module%install%", 100000)

    cursor.execute("""
        SELECT
            COALESCE(serial_no, '') AS serial_no,
            COALESCE(activity_name, '0') AS activity_name,
            COALESCE(uom, '') AS uom,
            COALESCE(vendor_name, '0') AS vendor_name,
            logged_date,
            COALESCE(MAX(scope), 0),
            COALESCE(MAX(
                CASE
                    WHEN previous_day_cumm = 0 AND cumulative_completion = 0 AND completion_percentage > 0 AND scope > 0
                    THEN GREATEST((scope * completion_percentage / 100) - today_progress, 0)
                    ELSE previous_day_cumm
                END
            ), 0),
            COALESCE(SUM(today_progress), 0),
            COALESCE(MAX(
                CASE
                    WHEN cumulative_completion = 0 AND completion_percentage > 0 AND scope > 0
                    THEN scope * completion_percentage / 100
                    ELSE cumulative_completion
                END
            ), 0),
            COALESCE(MAX(
                CASE
                    WHEN to_be_completed = 0 AND completion_percentage > 0 AND scope > 0
                    THEN GREATEST(scope - (scope * completion_percentage / 100), 0)
                    ELSE to_be_completed
                END
            ), 0),
            COALESCE(AVG(completion_percentage), 0),
            COALESCE(MAX(remarks), '')
        FROM running_progress_metrics
        GROUP BY serial_no, activity_name, uom, vendor_name, logged_date
        ORDER BY activity_name ASC, vendor_name ASC, logged_date ASC
    """)
    activity_rows = cursor.fetchall()

    activity_map = {}
    for a_row in activity_rows:
        act_name = a_row[1] or "0"
        vendor_name = a_row[3] or "0"
        logged_date = a_row[4]
        progress_value = float(a_row[7] or 0.0)
        key = (act_name or "0", vendor_name or "0")
        existing = activity_map.get(key)

        if existing is None:
            activity_map[key] = a_row
            continue

        existing_progress = float(existing[7] or 0.0)
        existing_date = existing[4]

        should_replace = False
        if progress_value > 0 and existing_progress <= 0:
            should_replace = True
        elif (progress_value > 0) == (existing_progress > 0):
            if logged_date and (not existing_date or logged_date > existing_date):
                should_replace = True

        if should_replace:
            activity_map[key] = a_row

    detailed_activity_table = []

    def table_sort_key(row):
        return (row[4] or datetime.min.date(), row[1] or "0", row[3] or "0")

    for a_row in sorted(activity_map.values(), key=table_sort_key, reverse=True):
        serial_no = a_row[0] or ""
        act_name = a_row[1] or "0"
        uom = a_row[2] or ""
        v_owner = a_row[3] or "0"
        dpr_date = a_row[4]
        t_scope = float(a_row[5] or 0.0)
        previous_val = float(a_row[6] or 0.0)
        t_prog = float(a_row[7] or 0.0)
        cumulative_val = float(a_row[8] or 0.0)
        to_be_completed_val = float(a_row[9] or 0.0)
        avg_comp = float(a_row[10] or 0.0)
        completion_display = round(avg_comp, 2)
        completion_width = max(0, min(completion_display, 100))
        remarks = a_row[11] or ""

        detailed_activity_table.append({
         "serial_no": serial_no,
         "activity": act_name,
         "uom": uom,
         "vendor": v_owner,
         "scope": f"{t_scope:,.2f}",
         "previous_day_cumm": f"{previous_val:,.2f}",
         "progress": f"{t_prog:,.2f}",
         "cumulative_completion": f"{cumulative_val:,.2f}",
         "to_be_completed": f"{to_be_completed_val:,.2f}",
         "completion": completion_display,
         "completion_width": completion_width,
         "remarks": remarks,
         "date": format_dpr_date(dpr_date)
        })

    return render_template(
        "progress_review.html",
        pile_casting_chart=pile_casting_chart,
        module_installation_chart=module_installation_chart,
        detailed_activity_table=detailed_activity_table
    )

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
